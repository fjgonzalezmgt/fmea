"""
Generador de FMEA usando OpenAI API.

Este módulo proporciona la clase FMEAGenerator que permite generar análisis
FMEA (Failure Mode and Effects Analysis) automáticamente utilizando modelos
de lenguaje de OpenAI.

Clases
------
FMEAGenerator
    Generador principal de análisis FMEA con integración a OpenAI

Notas
-----
El módulo utiliza plantillas de prompts externas almacenadas en archivos
markdown en la carpeta 'prompts/' para mayor flexibilidad y mantenibilidad.

Ejemplos
--------
>>> from fmea_generator import FMEAGenerator
>>> import pandas as pd
>>> 
>>> # Crear generador
>>> generator = FMEAGenerator(api_key="sk-...", language="es")
>>> 
>>> # Cargar datos del proceso
>>> df = pd.read_excel("proceso.xlsx")
>>> 
>>> # Generar FMEA
>>> fmea_df = generator.generate_fmea(df, num_failures=2)
>>> 
>>> # Exportar a Excel
>>> generator.export_to_excel(fmea_df, "fmea_resultado.xlsx")
"""
import pandas as pd
import json
import os
from typing import List, Dict, Optional
from openai import OpenAI


class FMEAGenerator:
    """
    Generador de análisis FMEA usando IA.
    
    Esta clase proporciona métodos para generar análisis FMEA (Failure Mode and Effects Analysis)
    automáticamente utilizando modelos de lenguaje de OpenAI.
    """
    
    def __init__(self, api_key: str, model: str = "gpt-5.4", language: str = "es"):
        """
        Inicializar el generador FMEA.
        
        Parameters
        ----------
        api_key : str
            API key de OpenAI para autenticación
        model : str, optional
            Modelo de OpenAI a utilizar (default: "gpt-5.4")
            Opciones: 'gpt-5.4', 'gpt-4o', 'gpt-4o-mini', 'gpt-4-turbo', 'gpt-3.5-turbo'
        language : str, optional
            Idioma de salida del análisis (default: "es")
            Opciones: 'es' (español), 'en' (inglés)
        
        Attributes
        ----------
        client : OpenAI
            Cliente de OpenAI para realizar llamadas a la API
        model : str
            Modelo configurado para el análisis
        language : str
            Idioma configurado para el análisis
        prompts_dir : str
            Ruta al directorio que contiene las plantillas de prompts
        """
        self.client = OpenAI(api_key=api_key)
        self.model = model
        self.language = language
        self.prompts_dir = os.path.join(os.path.dirname(__file__), 'prompts')
    
    def load_prompt_template(self, language: str) -> str:
        """
        Cargar plantilla de prompt desde archivo markdown.
        
        Parameters
        ----------
        language : str
            Código de idioma de la plantilla ('es' o 'en')
            
        Returns
        -------
        str
            Contenido de la plantilla de prompt
            
        Raises
        ------
        FileNotFoundError
            Si no se encuentra el archivo de plantilla para el idioma especificado
        Exception
            Si ocurre un error al leer el archivo
        """
        prompt_file = os.path.join(self.prompts_dir, f'fmea_prompt_{language}.md')
        
        try:
            with open(prompt_file, 'r', encoding='utf-8') as f:
                return f.read()
        except FileNotFoundError:
            raise FileNotFoundError(f"No se encontró el archivo de prompt: {prompt_file}")
        except Exception as e:
            raise Exception(f"Error al leer el archivo de prompt: {str(e)}")
        
    def read_process_excel(self, file) -> pd.DataFrame:
        """
        Leer archivo Excel con pasos del proceso.
        
        Parameters
        ----------
        file : str or BytesIO
            Archivo Excel (ruta del archivo o objeto BytesIO)
            
        Returns
        -------
        pd.DataFrame
            DataFrame con los pasos del proceso cargados desde el Excel
            
        Raises
        ------
        ValueError
            Si hay un error al leer o procesar el archivo Excel
        """
        try:
            df = pd.read_excel(file)
            return df
        except Exception as e:
            raise ValueError(f"Error al leer el archivo Excel: {str(e)}")
    
    def validate_process_data(self, df: pd.DataFrame) -> tuple[bool, str]:
        """
        Validar que el DataFrame tenga las columnas necesarias.
        
        Verifica que el DataFrame contenga al menos una columna de descripción
        del proceso con nombres comunes aceptados.
        
        Parameters
        ----------
        df : pd.DataFrame
            DataFrame a validar
            
        Returns
        -------
        tuple[bool, str]
            Tupla (es_valido, mensaje_error) donde:
            - es_valido: True si la validación es exitosa, False en caso contrario
            - mensaje_error: Mensaje descriptivo del error (vacío si es válido)
        """
        # Columnas requeridas (al menos una de estas debe existir para descripción)
        required_cols = ['descripcion', 'descripción', 'description', 'nombre', 'nombre del paso', 'paso']
        
        # Verificar que exista al menos una columna de descripción
        has_description = any(col.lower() in [c.lower() for c in df.columns] for col in required_cols)
        
        if not has_description:
            return False, f"El Excel debe contener al menos una columna de descripción del paso. Columnas encontradas: {', '.join(df.columns)}"
        
        if len(df) == 0:
            return False, "El archivo Excel está vacío"
            
        return True, ""
    
    def get_description_column(self, df: pd.DataFrame) -> str:
        """
        Identificar la columna de descripción del proceso.
        
        Busca en el DataFrame una columna que contenga la descripción de los pasos
        del proceso, probando nombres comunes en español e inglés.
        
        Parameters
        ----------
        df : pd.DataFrame
            DataFrame con los datos del proceso
            
        Returns
        -------
        str
            Nombre de la columna identificada como descripción del proceso
            
        Notes
        -----
        Si no se encuentra una columna con nombre estándar, se utiliza la segunda
        columna (asumiendo que la primera es el número) o la primera columna disponible.
        """
        possible_names = ['descripcion', 'descripción', 'description', 'nombre', 'nombre del paso', 'paso', 'actividad']
        
        for col in df.columns:
            if col.lower() in possible_names:
                return col
                
        # Si no encuentra, usar la segunda columna (asumiendo que la primera es número)
        if len(df.columns) > 1:
            return df.columns[1]
        return df.columns[0]
    
    def get_step_info(self, row: pd.Series, desc_col: str) -> str:
        """
        Construir descripción completa del paso del proceso.
        
        Combina la descripción principal con información adicional disponible
        (responsable, entradas, salidas, recursos, etc.) para proporcionar
        contexto completo a la IA.
        
        Parameters
        ----------
        row : pd.Series
            Fila del DataFrame que representa un paso del proceso
        desc_col : str
            Nombre de la columna que contiene la descripción principal
            
        Returns
        -------
        str
            Descripción completa del paso formateada con toda la información disponible,
            separada por pipe (|)
        """
        info_parts = []
        
        # Descripción principal
        info_parts.append(f"Paso: {row[desc_col]}")
        
        # Agregar otras columnas relevantes si existen
        optional_cols = {
            'responsable': 'Responsable',
            'entradas': 'Entradas',
            'entrada': 'Entradas',
            'salidas': 'Salidas',
            'salida': 'Salidas',
            'recursos': 'Recursos',
            'herramientas': 'Herramientas',
            'procedimiento': 'Procedimiento',
            'detalles': 'Detalles'
        }
        
        for col in row.index:
            col_lower = col.lower()
            if col_lower in optional_cols and pd.notna(row[col]) and str(row[col]).strip():
                info_parts.append(f"{optional_cols[col_lower]}: {row[col]}")
        
        return " | ".join(info_parts)
    
    def build_fmea_prompt(self, process_steps: List[str], num_failures: int = 2) -> str:
        """
        Construir el prompt para generar análisis FMEA.
        
        Carga la plantilla de prompt correspondiente al idioma configurado,
        formatea la lista de pasos del proceso y reemplaza los placeholders.
        
        Parameters
        ----------
        process_steps : List[str]
            Lista de descripciones de pasos del proceso a analizar
        num_failures : int, optional
            Número de modos de fallo a generar por cada paso (default: 2)
            
        Returns
        -------
        str
            Prompt completo formateado para enviar a la API de OpenAI
            
        Notes
        -----
        Utiliza placeholders en la plantilla:
        - {num_failures}: Reemplazado por el número de modos de fallo
        - {process_steps}: Reemplazado por la lista enumerada de pasos
        """
        # Cargar template desde archivo
        template = self.load_prompt_template(self.language)
        
        # Formatear lista de pasos del proceso
        process_steps_text = '\n'.join(f"{i+1}. {step}" for i, step in enumerate(process_steps))
        
        # Reemplazar placeholders
        prompt = template.replace('{num_failures}', str(num_failures))
        prompt = prompt.replace('{process_steps}', process_steps_text)
        
        return prompt
    
    def generate_fmea(self, df: pd.DataFrame, num_failures: int = 2, progress_callback=None) -> pd.DataFrame:
        """
        Generar análisis FMEA completo.
        
        Procesa el DataFrame de pasos del proceso, construye el prompt,
        llama a la API de OpenAI, parsea la respuesta y retorna el análisis
        FMEA estructurado con cálculo de RPN.
        
        Parameters
        ----------
        df : pd.DataFrame
            DataFrame con los pasos del proceso a analizar
        num_failures : int, optional
            Número de modos de fallo a generar por paso (default: 2)
        progress_callback : callable, optional
            Función de callback para reportar progreso.
            Firma: callback(progress: float, message: str)
            donde progress es un valor entre 0 y 1
            
        Returns
        -------
        pd.DataFrame
            DataFrame con el análisis FMEA completo incluyendo:
            - numero_paso: Número del paso del proceso
            - paso_proceso: Descripción del paso
            - modo_fallo: Modo de fallo identificado
            - efecto_fallo: Efecto del modo de fallo
            - severidad: Calificación de severidad (1-10)
            - causa_potencial: Causa potencial del fallo
            - ocurrencia: Calificación de ocurrencia (1-10)
            - controles_actuales: Controles actuales del proceso
            - deteccion: Calificación de detección (1-10)
            - RPN: Risk Priority Number (S × O × D)
            - acciones_recomendadas: Acciones recomendadas
            
        Raises
        ------
        ValueError
            Si los datos del proceso no son válidos, si hay error al parsear
            la respuesta de OpenAI, o si ocurre un error durante la generación
        """
        # Validar datos
        is_valid, error_msg = self.validate_process_data(df)
        if not is_valid:
            raise ValueError(error_msg)
        
        # Identificar columna de descripción
        desc_col = self.get_description_column(df)
        
        # Construir lista de pasos con toda la información
        process_steps = []
        for idx, row in df.iterrows():
            step_info = self.get_step_info(row, desc_col)
            process_steps.append(step_info)
        
        # Construir prompt
        prompt = self.build_fmea_prompt(process_steps, num_failures)
        
        if progress_callback:
            progress_callback(0.3, "Enviando solicitud a OpenAI...")
        
        # Llamar a OpenAI API
        try:
            response = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {"role": "system", "content": "Eres un experto en análisis FMEA y calidad de procesos. Siempre respondes en formato JSON válido."},
                    {"role": "user", "content": prompt}
                ],
                response_format={"type": "json_object"},
                temperature=0.7
            )
            
            if progress_callback:
                progress_callback(0.7, "Procesando respuesta...")
            
            # Parsear respuesta
            result = json.loads(response.choices[0].message.content)
            
            # Convertir a DataFrame
            fmea_df = pd.DataFrame(result['fmea_items'])
            
            # Calcular RPN (Risk Priority Number)
            fmea_df['RPN'] = fmea_df['severidad'] * fmea_df['ocurrencia'] * fmea_df['deteccion']
            
            # Reordenar columnas
            column_order = [
                'numero_paso', 'paso_proceso', 'modo_fallo', 'efecto_fallo',
                'severidad', 'causa_potencial', 'ocurrencia', 'controles_actuales',
                'deteccion', 'RPN', 'acciones_recomendadas'
            ]
            
            # Solo reordenar las columnas que existen
            existing_cols = [col for col in column_order if col in fmea_df.columns]
            fmea_df = fmea_df[existing_cols]
            
            if progress_callback:
                progress_callback(1.0, "¡FMEA generado exitosamente!")
            
            return fmea_df
            
        except json.JSONDecodeError as e:
            raise ValueError(f"Error al parsear la respuesta de OpenAI: {str(e)}")
        except Exception as e:
            raise ValueError(f"Error al generar FMEA: {str(e)}")
    
    def export_to_excel(self, fmea_df: pd.DataFrame, output_file: str):
        """
        Exportar análisis FMEA a archivo Excel con formato profesional.
        
        Crea un archivo Excel con formato condicional basado en valores de RPN,
        encabezados formateados, filtros automáticos y columnas ajustadas.
        
        Parameters
        ----------
        fmea_df : pd.DataFrame
            DataFrame con el análisis FMEA a exportar
        output_file : str
            Ruta del archivo de salida Excel (.xlsx)
            
        Notes
        -----
        Aplicaciones de formato:
        - Encabezados: Fondo azul oscuro con texto blanco en negrita
        - RPN < 50: Fondo verde (riesgo bajo)
        - 50 ≤ RPN ≤ 100: Fondo amarillo (riesgo medio)
        - RPN > 100: Fondo rojo (riesgo alto)
        - Filtros automáticos habilitados en todas las columnas
        - Ancho de columnas ajustado al contenido (máximo 50 caracteres)
        """
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Exportar a Excel
        fmea_df.to_excel(output_file, index=False, sheet_name='FMEA')
        
        # Cargar workbook para aplicar formato
        wb = load_workbook(output_file)
        ws = wb['FMEA']
        
        # Formato de encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Formato condicional para RPN
        rpn_col_idx = None
        for idx, col in enumerate(fmea_df.columns, 1):
            if col == 'RPN':
                rpn_col_idx = idx
                break
        
        if rpn_col_idx:
            col_letter = get_column_letter(rpn_col_idx)
            
            # Verde para RPN < 50
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            # Amarillo para RPN 50-100
            yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            # Rojo para RPN > 100
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            for row_idx in range(2, len(fmea_df) + 2):
                cell = ws[f"{col_letter}{row_idx}"]
                rpn_value = cell.value
                
                if rpn_value is not None:
                    if rpn_value > 100:
                        cell.fill = red_fill
                    elif rpn_value >= 50:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = green_fill
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Habilitar filtros
        ws.auto_filter.ref = ws.dimensions
        
        # Guardar
        wb.save(output_file)
